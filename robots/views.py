from django.http import HttpResponse
import csv
from openpyxl import Workbook
from .models import Robot
from datetime import datetime, timedelta

# Создание робота с полями модели, версии и даты создания
def create_robot(model, version, created):
    return Robot.objects.create(
        model=model,
        version=version,
        created=datetime.strptime(created, '%Y-%m-%d %H:%M:%S')
    )

# Функция для фильтрации роботов по дате создания за последнюю неделю и сортировки по дате
def filter_and_sort_robots_by_data():
    # Начало недели
    today = datetime.now()
    start_of_week = today - timedelta(days=today.weekday())

    # Фильтрация роботов по дате создания за последнюю неделю
    return Robot.objects.filter(created__gte=start_of_week).order_by('created')

# Функция для сортировки роботов по моделям с версиями
def sort_robots_by_model_and_version(robots):
    # Словарь для хранения данных по моделям и версиям
    robot_data = {}

    # Заполнение словаря данными
    for robot in robots:
        model = robot.model
        version = robot.version
        count = robot_data.get((model, version), 0)
        robot_data[(model, version)] = count + 1

    return robot_data

# Функция для создания Excel-файла
def generate_robot_summary_excel(request):
    # Создание Excel файла
    workbook = Workbook()
    
    # Фильтруем и сортируем роботов
    robots = filter_and_sort_robots_by_data()
    
    # Получаем список уникальных моделей
    unique_models = set(robots.values_list('model', flat=True))
    
    # Создаем отдельный лист в Excel для каждой модели
    for model in unique_models:
        sheet = workbook.create_sheet(title=model)
        
        # Заголовки столбцов
        sheet['A1'] = "Модель"
        sheet['B1'] = "Версия"
        sheet['C1'] = "Количество за неделю"
        
        # Фильтруем роботов для текущей модели
        model_robots = robots.filter(model=model)
        
        # Создаем словарь с информацией о роботах для текущей модели
        robot_data = sort_robots_by_version(model_robots)
        
        # Заполняем таблицу данными из словаря
        row_num = 2  # Начнем с второй строки, так как первая строка занята заголовками
        for (version, count) in robot_data.items():
            sheet.cell(row=row_num, column=1, value=model)
            sheet.cell(row=row_num, column=2, value=version)
            sheet.cell(row=row_num, column=3, value=count)
            row_num += 1
    
    # Удаляем первый лист (пустой лист по умолчанию)
    del workbook['Sheet']
    
    # Создаем HTTP-ответ с файлом Excel
    response = HttpResponse(content_type='application/ms-excel')
    response['Content-Disposition'] = 'attachment; filename="robot_summary.xlsx"'

    # Сохраняем файл Excel в HTTP-ответ
    workbook.save(response)

    return response

# Функция для сортировки роботов по версиям
def sort_robots_by_version(robots):
    # Создаем словарь для хранения данных по версиям
    robot_data = {}

    # Заполняем словарь данными
    for robot in robots:
        version = robot.version
        count = robot_data.get(version, 0)
        robot_data[version] = count + 1

    return robot_data

# Функция для создания CSV-файла
def generate_robot_summary_csv(request):
    # Фильтруем и сортируем роботов
    robots = filter_and_sort_robots_by_data()

    # Создаем CSV-файл
    response = HttpResponse(content_type='text/csv')
    response['Content-Disposition'] = 'attachment; filename="robot_summary.csv"'

    writer = csv.writer(response)
    writer.writerow(['Модель', 'Версия', 'Количество за неделю'])

    # Создаем словарь с информацией о роботах
    robot_data = sort_robots_by_model_and_version(robots)

    # Заполняем CSV данными из словаря
    for (model, version), count in robot_data.items():
        writer.writerow([model, version, count])

    return response
